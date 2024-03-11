# open_workbook.py

from openpyxl import load_workbook

def show_methods(obj):
    # methods_and_props_without_specjal = [attr for attr in dir(obj) if not attr.startswith('__')]
    methods_and_props_without_specjal_and_private = [f"{ 'Method   >' if callable(getattr(obj, attr)) else 'Atribute >'} {attr}" for attr in dir(obj) if not (attr.startswith('__') or attr.startswith('_'))]
    print(*sorted(methods_and_props_without_specjal_and_private), sep="\n")

def open_workbook(path):
    workbook = load_workbook(filename=path)
    print(f"Worksheet names: {workbook.sheetnames}")
    sheet = workbook.active
    print(sheet)
    print(f"The title of the Worksheet is: {sheet.title}")
    """
    Worksheet names: ['Sheet 1 - Books', 'Sales']
    <Worksheet "Sales">
    The title of the Worksheet is: Sales
    """

    print(show_methods(workbook))
    """
    Atribute > active
Atribute > calculation
Atribute > chartsheets
Atribute > code_name
Atribute > custom_doc_props
Atribute > data_only
Atribute > defined_names
Atribute > encoding
Atribute > epoch
Atribute > excel_base_date
Atribute > is_template
Atribute > iso_dates
Atribute > loaded_theme
Atribute > mime_type
Atribute > named_styles
Atribute > path
Atribute > properties
Atribute > read_only
Atribute > rels
Atribute > security
Atribute > shared_strings
Atribute > sheetnames
Atribute > style_names
Atribute > template
Atribute > vba_archive
Atribute > views
Atribute > worksheets
Atribute > write_only
Method   > add_named_style
Method   > close
Method   > copy_worksheet
Method   > create_chartsheet
Method   > create_named_range
Method   > create_sheet
Method   > get_index
Method   > get_sheet_by_name
Method   > get_sheet_names
Method   > index
Method   > move_sheet
Method   > remove
Method   > remove_sheet
Method   > save
"""
    print(dir(sheet))
    """
    ['BREAK_COLUMN', 'BREAK_NONE', 'BREAK_ROW', 'HeaderFooter', 'ORIENTATION_LANDSCAPE', 'ORIENTATION_PORTRAIT', 'PAPERSIZE_A3', 'PAPERSIZE_A4', 'PAPERSIZE_A4_SMALL', 'PAPERSIZE_A5', 'PAPERSIZE_EXECUTIVE', 'PAPERSIZE_LEDGER', 'PAPERSIZE_LEGAL', 'PAPERSIZE_LETTER', 'PAPERSIZE_LETTER_SMALL', 'PAPERSIZE_STATEMENT', 'PAPERSIZE_TABLOID', 'SHEETSTATE_HIDDEN', 'SHEETSTATE_VERYHIDDEN', 'SHEETSTATE_VISIBLE', '_WorkbookChild__title', '__class__', '__delattr__', '__delitem__', '__dict__', '__dir__', '__doc__', '__eq__', '__format__', '__ge__', '__getattribute__', '__getitem__', '__getstate__', '__gt__', '__hash__', '__init__', '__init_subclass__', '__iter__', '__le__', '__lt__', '__module__', '__ne__', '__new__', '__reduce__', '__reduce_ex__', '__repr__', '__setattr__', '__setitem__', '__sizeof__', '__str__', '__subclasshook__', '__weakref__', '_add_cell', '_add_column', '_add_row', '_cells', '_cells_by_col', '_cells_by_row', '_charts', '_clean_merge_range', '_comments', '_current_row', '_default_title', '_drawing', '_get_cell', '_hyperlinks', '_id', '_images', '_invalid_row', '_move_cell', '_move_cells', '_parent', '_path', '_pivots', '_print_area', '_print_cols', '_print_rows', '_rel_type', '_rels', '_setup', '_tables', 'active_cell', 'add_chart', 'add_data_validation', 'add_image', 'add_pivot', 'add_table', 'append', 'array_formulae', 'auto_filter', 'calculate_dimension', 'cell', 'col_breaks', 'column_dimensions', 'columns', 'conditional_formatting', 'data_validations', 'defined_names', 'delete_cols', 'delete_rows', 'dimensions', 'encoding', 'evenFooter', 'evenHeader', 'firstFooter', 'firstHeader', 'freeze_panes', 'insert_cols', 'insert_rows', 'iter_cols', 'iter_rows', 'legacy_drawing', 'max_column', 'max_row', 'merge_cells', 'merged_cell_ranges', 'merged_cells', 'mime_type', 'min_column', 'min_row', 'move_range', 'oddFooter', 'oddHeader', 'page_margins', 'page_setup', 'parent', 'path', 'print_area', 'print_options', 'print_title_cols', 'print_title_rows', 'print_titles', 'protection', 'row_breaks', 'row_dimensions', 'rows', 'scenarios', 'selected_cell', 'set_printer_settings', 'sheet_format', 'sheet_properties', 'sheet_state', 'sheet_view', 'show_gridlines', 'tables', 'title', 'unmerge_cells', 'values', 'views']
    """
    # print(show_methods(sheet))

if __name__ == "__main__":
    open_workbook("books.xlsx")
