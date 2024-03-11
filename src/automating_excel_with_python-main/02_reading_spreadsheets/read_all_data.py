# read_all_data.py

import openpyxl
from openpyxl import load_workbook


def read_all_data(path):
    workbook = load_workbook(filename=path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"{sheet.title=}")
        for row in sheet.rows:
            for cell in row:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    # Skip this cell - jeśli komórka jest scalona z kilku komórek
                    continue

                print(f"{cell.column_letter}{cell.row} = {cell.value}")

    """
sheet.title='Sheet 1 - Books'
A1 = Books
A2 = Title
B2 = Author
C2 = Publisher
D2 = Publishing Date
E2 = ISBN
F2 = None
G2 = None
A3 = Python 101
B3 = Mike Driscoll
C3 = Mouse vs Python
D3 = 2020
E3 = 1234567890
F3 = None
G3 = None
A4 = wxPython Recipes
B4 = Mike Driscoll
C4 = Apress
D4 = 2018
E4 = 978-1-4842-3237-8
F4 = None
G4 = None
A5 = Python Interviews
B5 = Mike Driscoll
C5 = Packt Publishing
D5 = 2018
E5 = 9781788399081
F5 = None
G5 = None
A6 = None
B6 = None
C6 = None
D6 = None
E6 = None
F6 = None
G6 = None
A7 = None
B7 = None
C7 = None
D7 = None
E7 = None
F7 = None
G7 = None
A8 = None
B8 = None
C8 = None
D8 = None
E8 = None
F8 = None
G8 = None
A9 = None
B9 = None
C9 = None
D9 = None
E9 = None
F9 = None
G9 = None
A10 = None
B10 = None
C10 = None
D10 = None
E10 = None
F10 = None
G10 = None
A11 = None
B11 = None
C11 = None
D11 = None
E11 = None
F11 = None
G11 = None
A12 = None
B12 = None
C12 = None
D12 = None
E12 = None
F12 = None
G12 = None
A13 = None
B13 = None
C13 = None
D13 = None
E13 = None
F13 = None
G13 = None
A14 = None
B14 = None
C14 = None
D14 = None
E14 = None
F14 = None
G14 = None
A15 = None
B15 = None
C15 = None
D15 = None
E15 = None
F15 = None
G15 = None
A16 = None
B16 = None
C16 = None
D16 = None
E16 = None
F16 = None
G16 = None
A17 = None
B17 = None
C17 = None
D17 = None
E17 = None
F17 = None
G17 = None
A18 = None
B18 = None
C18 = None
D18 = None
E18 = None
F18 = None
G18 = None
A19 = None
B19 = None
C19 = None
D19 = None
E19 = None
F19 = None
G19 = None
A20 = None
B20 = None
C20 = None
D20 = None
E20 = None
F20 = None
G20 = None
A21 = None
B21 = None
C21 = None
D21 = None
E21 = None
F21 = None
G21 = None
A22 = None
B22 = None
C22 = None
D22 = None
E22 = None
F22 = None
G22 = None
A23 = None
B23 = None
C23 = None
D23 = None
E23 = None
F23 = None
G23 = None
sheet.title='Sales'
A1 = Title
B1 = Amazon
C1 = Leanpub
D1 = Gumroad
A2 = Python 101
B2 = 100
C2 = 433
D2 = 10
A3 = wxPython Recipes
B3 = 5
C3 = 0
D3 = 0
A4 = Python Interviews
B4 = 10
C4 = 0
D4 = 0    
    """
if __name__ == "__main__":
    read_all_data("books.xlsx")