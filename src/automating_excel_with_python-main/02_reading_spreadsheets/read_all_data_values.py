# read_all_data_values.py

import openpyxl
from openpyxl import load_workbook


def read_all_data(path):
    workbook = load_workbook(filename=path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"{sheet.title=}")
        for value in sheet.iter_rows(values_only=True):
            print(value)
"""
używa iteratora po wierszach

sheet.title='Sheet 1 - Books'
('Books', None, None, None, None, None, None)
('Title', 'Author', 'Publisher', 'Publishing Date', 'ISBN', None, None)
('Python 101', 'Mike Driscoll', 'Mouse vs Python', 2020, 1234567890, None, None)
('wxPython Recipes', 'Mike Driscoll', 'Apress', 2018, '978-1-4842-3237-8', None, None)
('Python Interviews', 'Mike Driscoll', 'Packt Publishing', 2018, 9781788399081, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
(None, None, None, None, None, None, None)
sheet.title='Sales'
('Title', 'Amazon', 'Leanpub', 'Gumroad')
('Python 101', 100, 433, 10)
('wxPython Recipes', 5, 0, 0)
('Python Interviews', 10, 0, 0)
"""

if __name__ == "__main__":
    read_all_data("books.xlsx")