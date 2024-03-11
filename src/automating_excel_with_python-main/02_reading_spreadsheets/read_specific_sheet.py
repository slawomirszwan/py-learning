# read_specific_sheet.py

from openpyxl import load_workbook


def open_workbook(path, sheet_name):
    try:
        workbook = load_workbook(filename=path)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"The title of the Worksheet is: {sheet.title}")
            print(f"Cells that contain data: {sheet.calculate_dimension()}")
        else:
            print(f"Error: Workbook don't have sheet {sheet_name}")
    except FileNotFoundError as e:
        print(f"Error: File not found {path}")

if __name__ == "__main__":
    open_workbook("books1.xlsx", sheet_name="Sales")
