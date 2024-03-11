"""
kod kolegi z forum
"""
# [1] list comprehesnion
# [2] decorators
# [3] unique values using set
# [4] unused variable
# [5] unpacking operation
# [6] oneliner if

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import pandas as pd

from funkcje import poprawny_pesel, pokaz_zapis

# import plików do DataFrame
lekarze = pd.read_csv('dane/lekarze.txt', sep='\t', index_col=0)
pacjenci = pd.read_csv('dane/pacjenci.txt', sep='\t', index_col=0)
wizyty = pd.read_csv('dane/wizyty.txt', sep='\t', index_col=0)

# definicje kolorów i stylu linii
CZERWONY = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
POMARANCZOWY = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')
ZIELONY = PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid')
SZARY = PatternFill(start_color='AAAAAA', end_color='AAAAAA', fill_type='solid')
JASNOSZARY = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
LINIA = Side(border_style="thin", color="000000")


def ustaw_styl_komorek(ws, nr_wiersza, zakres, kolor, bold=False, italic=False):
    """
    ustwawia kolor tła komórek, styl znaków i obramowanie (openpyxl)
    :param ws: worshFFt
    :param nr_wiersza: numer wiersza
    :param zakres: ilość kolumn
    :param kolor: PaternFill kolor
    :param bold: czy font ma byc pogrubiony
    :param italic: czy font ma byc pochylony
    """
    for i in range(zakres):
        komorka = ws.cell(row=nr_wiersza, column=i + 1)
        komorka.fill = kolor
        komorka.font = Font(bold=bold, italic=italic)
        komorka.border = Border(top=LINIA, left=LINIA, right=LINIA, bottom=LINIA)


def ustaw_szerokosc_kolumn(ws):
    """
    ustawia automatycznie szerokość kolumn
    :param ws: arkusz (worshFFt) na którym działa
    """
    # stała określająca ile znaków więcej ma mieć komórka
    NADMIAR = 2

    for komorki in ws.columns:
        znakow = 0
        # pobiera literę kolumny Excel
        nazwa = komorki[0].column_letter
        # wylicza szekorkość kolumny
        for komorka in komorki:
            if len(str(komorka.value)) > znakow:
                znakow = len(str(komorka.value))
        szerokosc = znakow + NADMIAR
        # ustawia szerokość kolumny
        ws.column_dimensions[nazwa].width = szerokosc

    # [2]


@pokaz_zapis
def zadanie_8():
    """
    Przygotuj zestawienie i zapisz je do pliku excel.
    Zestawienie powinno zawierać listę pacjentów i liczbę ich wizyt.
    Pokoloruj za pomocą python 3 (?) pacjentów najczęściej korzystający z wizyt:
        1 miejsce - czerwony
        2 miejsce - pomarańczowy
        3 miejsce - zielony
    """
    global pacjenci, wizyty

    # otwórz plik excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Pacjenci i ich wizyty"

    # w pandas zrób zestawienie pacjenci i wizyty
    id_pacjenta_wizyt = wizyty.groupby(['Id_pacjenta']).count(). \
        set_axis(['Ilosc_wizyt'], axis='columns')
    wynik = pd.merge(pacjenci, id_pacjenta_wizyt, on='Id_pacjenta')
    # set (unikalne wartości) z liczbami wizyt, posortowane malejąco
    # [3]
    ilosci_wizyt = sorted(set(wynik['Ilosc_wizyt']), reverse=True)

    # stwórz nazwy kolumn w szarym tle i z pogrubionymi literami
    ws.append(['Nazwisko', 'Imie', 'Ilosc_wizyt'])
    ustaw_styl_komorek(ws, 1, 3, SZARY, bold=True)

    # dla każdego pacjenta
    # [4]
    for _, row in wynik.iterrows():
        # dodaj wiersz do pliku excel
        ws.append([row['Nazwisko'], row['Imie'], row['Ilosc_wizyt']])

        # pobierz nr wiersza
        nr_wiersza = ws._current_row

        # Piękna mielonka! Cudowna mielonka! W zależności od ilości wizyt ustaw kolor
        # 1 miejsce
        if row['Ilosc_wizyt'] == ilosci_wizyt[0]:
            ustaw_styl_komorek(ws, nr_wiersza, 3, CZERWONY)
        # 2 miejsce
        elif row['Ilosc_wizyt'] == ilosci_wizyt[1]:
            ustaw_styl_komorek(ws, nr_wiersza, 3, POMARANCZOWY)
        # 3 miejsce
        elif row['Ilosc_wizyt'] == ilosci_wizyt[2]:
            ustaw_styl_komorek(ws, nr_wiersza, 3, ZIELONY)
        else:
            ustaw_styl_komorek(ws, nr_wiersza, 3, JASNOSZARY)

    ustaw_szerokosc_kolumn(ws)
    # zapisz plik excel
    wb.save("dane/zadanie_8.xlsx")


# [2]
@pokaz_zapis
def zadanie_9():
    """
    Wygeneruj raport błędnych numerów pesel lekarzy.
    Raport składa się z całej listy lekarzy, ale błędne pesele są na czerwono
    """
    global lekarze

    # zamień PESEL na string
    lekarze['PESEL'] = lekarze['PESEL'].astype(str)

    # otwórz plik excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Lekarze"

    # A teraz coś zupełnie innego
    #  by nie używać stałych jak w poprzednim zadaniu (uniwersalne) pobrać:
    # - ilość kolumn
    kolumn = len(lekarze.columns)
    # - nazwy kolumn z dataframe
    nazwy = lekarze.columns.tolist()
    # i zapisz w 1 kolumnie w szarym tle i z pogrubionymi literami
    ws.append(nazwy)
    ustaw_styl_komorek(ws, 1, kolumn, SZARY, bold=True)

    # dodawaj wiersze i pokoloruj używając fukcji sprawdz_pesel
    # [4]
    for _, row in lekarze.iterrows():
        # zrób listę z kolumnami i dodaj wiersz do pliku excel
        # [1]
        ws.append([row[nazwy[i]] for i in range(kolumn)])
        # pobierz nr wiersza
        nr_wiersza = ws._current_row

        ustaw_styl_komorek(ws, nr_wiersza, kolumn, JASNOSZARY)
        # namierz komórkę z PESELem
        k = ws.cell(row=nr_wiersza, column=nazwy.index('PESEL') + 1)
        # Na koniec mój panie, listek miętowy.
        # [5] [6]
        k.font = Font(color='FF0000' if poprawny_pesel(*row['PESEL']) else '000000')

    ustaw_szerokosc_kolumn(ws)
    # zapisz plik excel
    wb.save("dane/zadanie_9.xlsx")


if __name__ == "__main__":
    zadanie_8()
    zadanie_9()
