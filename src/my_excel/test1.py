from openpyxl import Workbook

wb = Workbook()

ws = wb.active

"""
>>> ws1 = wb.create_sheet("Mysheet") # insert at the end (default)
# or
>>> ws2 = wb.create_sheet("Mysheet", 0) # insert at first position
# or
>>> ws3 = wb.create_sheet("Mysheet", -1) # insert at the penultimate position
"""


ws.title = "Sheet"


for sheet in wb:
    print(sheet.title)

"""
>>> source = wb.active
>>> target = wb.copy_worksheet(source)
"""

# print(*dir(wb), sep="\n")
"""
_Workbook__write_only
__class__
__contains__
__delattr__
__delitem__
__dict__
__dir__
__doc__
__eq__
__format__
__ge__
__getattribute__
__getitem__
__getstate__
__gt__
__hash__
__init__
__init_subclass__
__iter__
__le__
__lt__
__module__
__ne__
__new__
__reduce__
__reduce_ex__
__repr__
__setattr__
__sizeof__
__str__
__subclasshook__
__weakref__
_active_sheet_index
_add_sheet
_alignments
_borders
_cell_styles
_colors
_data_only
_date_formats
_differential_styles
_duplicate_name
_epoch
_external_links
_fills
_fonts
_named_styles
_number_formats
_pivots
_protections
_read_only
_setup_styles
_sheets
_table_styles
_timedelta_formats

active
add_named_style
calculation
chartsheets
close
code_name
copy_worksheet
create_chartsheet
create_named_range
create_sheet
custom_doc_props
data_only
defined_names
encoding
epoch
excel_base_date
get_index
get_sheet_by_name
get_sheet_names
index
is_template
iso_dates
loaded_theme
mime_type
move_sheet
named_styles
path
properties
read_only
rels
remove
remove_sheet
save
security
shared_strings
sheetnames
style_names
template
vba_archive
views
worksheets
write_only
"""

print(wb.sheetnames)
# ['Sheet']

# -------------------------------------------------
# Cell

# odwołanie do celi po oznaczeniu excel
c1 = ws['A4']

c1 = 4


# odwołanie do numeru wirsza i kolumny
c2 = ws.cell(row=4, column=2, value=10)


# --------------------------------
# Przykładowe dane
for row in range(1,20):
    for col in range(1,10):
        ws.cell(row=row, column=col, value="Hay")


# zakres cell
cell_range = ws['A1':'C2']

"""
>>> colC = ws['C']
>>> col_range = ws['C:D']
>>> row10 = ws[10]
>>> row_range = ws[5:10]
"""



"""
użycie iteracji

>>> for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
...    for cell in row:
...        print(cell)
<Cell Sheet1.A1>
<Cell Sheet1.B1>
<Cell Sheet1.C1>
<Cell Sheet1.A2>
<Cell Sheet1.B2>
<Cell Sheet1.C2>


>>> for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
...     for cell in col:
...         print(cell)
<Cell Sheet1.A1>
<Cell Sheet1.A2>
<Cell Sheet1.B1>
<Cell Sheet1.B2>
<Cell Sheet1.C1>
<Cell Sheet1.C2>

"""


# print(*dir(ws),sep="\n")
"""
BREAK_COLUMN
BREAK_NONE
BREAK_ROW
HeaderFooter
ORIENTATION_LANDSCAPE
ORIENTATION_PORTRAIT
PAPERSIZE_A3
PAPERSIZE_A4
PAPERSIZE_A4_SMALL
PAPERSIZE_A5
PAPERSIZE_EXECUTIVE
PAPERSIZE_LEDGER
PAPERSIZE_LEGAL
PAPERSIZE_LETTER
PAPERSIZE_LETTER_SMALL
PAPERSIZE_STATEMENT
PAPERSIZE_TABLOID
SHEETSTATE_HIDDEN
SHEETSTATE_VERYHIDDEN
SHEETSTATE_VISIBLE
_WorkbookChild__title
__class__
__delattr__
__delitem__
__dict__
__dir__
__doc__
__eq__
__format__
__ge__
__getattribute__
__getitem__
__getstate__
__gt__
__hash__
__init__
__init_subclass__
__iter__
__le__
__lt__
__module__
__ne__
__new__
__reduce__
__reduce_ex__
__repr__
__setattr__
__setitem__
__sizeof__
__str__
__subclasshook__
__weakref__
_add_cell
_add_column
_add_row
_cells
_cells_by_col
_cells_by_row
_charts
_clean_merge_range
_comments
_current_row
_default_title
_drawing
_get_cell
_hyperlinks
_id
_images
_invalid_row
_move_cell
_move_cells
_parent
_path
_pivots
_print_area
_print_cols
_print_rows
_rel_type
_rels
_setup
_tables
active_cell
add_chart
add_data_validation
add_image
add_pivot
add_table
append
array_formulae
auto_filter
calculate_dimension
cell
col_breaks
column_dimensions
columns
conditional_formatting
data_validations
defined_names
delete_cols
delete_rows
dimensions
encoding
evenFooter
evenHeader
firstFooter
firstHeader
freeze_panes
insert_cols
insert_rows
iter_cols
iter_rows
legacy_drawing
max_column
max_row
merge_cells
merged_cell_ranges
merged_cells
mime_type
min_column
min_row
move_range
oddFooter
oddHeader
page_margins
page_setup
parent
path
print_area
print_options
print_title_cols
print_title_rows
print_titles
protection
row_breaks
row_dimensions
rows
scenarios
selected_cell
set_printer_settings
sheet_format
sheet_properties
sheet_state
sheet_view
show_gridlines
tables
title
unmerge_cells
values
views
"""



"""
jak potrzeba iteracji po wszystkicj wierszach można użyć

ws.rows


>>> ws = wb.active
>>> ws['C9'] = 'hello world'
>>> tuple(ws.rows)
((<Cell Sheet.A1>, <Cell Sheet.B1>, <Cell Sheet.C1>),
(<Cell Sheet.A2>, <Cell Sheet.B2>, <Cell Sheet.C2>),
(<Cell Sheet.A3>, <Cell Sheet.B3>, <Cell Sheet.C3>),
(<Cell Sheet.A4>, <Cell Sheet.B4>, <Cell Sheet.C4>),
(<Cell Sheet.A5>, <Cell Sheet.B5>, <Cell Sheet.C5>),
(<Cell Sheet.A6>, <Cell Sheet.B6>, <Cell Sheet.C6>),
(<Cell Sheet.A7>, <Cell Sheet.B7>, <Cell Sheet.C7>),
(<Cell Sheet.A8>, <Cell Sheet.B8>, <Cell Sheet.C8>),
(<Cell Sheet.A9>, <Cell Sheet.B9>, <Cell Sheet.C9>))


albo  ws.columns

>>> tuple(ws.columns)
((<Cell Sheet.A1>,
<Cell Sheet.A2>,
<Cell Sheet.A3>,
<Cell Sheet.A4>,
<Cell Sheet.A5>,
<Cell Sheet.A6>,
...
<Cell Sheet.B7>,
<Cell Sheet.B8>,
<Cell Sheet.B9>),
(<Cell Sheet.C1>,
<Cell Sheet.C2>,
<Cell Sheet.C3>,
<Cell Sheet.C4>,
<Cell Sheet.C5>,
<Cell Sheet.C6>,
<Cell Sheet.C7>,
<Cell Sheet.C8>,
<Cell Sheet.C9>))

"""



"""
gdy chcemy tylko wartości z ws

ws.values

for row in ws.values:
   for value in row:
     print(value)
     
     
     


Both Worksheet.iter_rows() and Worksheet.iter_cols() can take the values_only parameter to return just the cell’s value:

iteratory maja parametr aby zwracały tylko wartości
values_only

>>> for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
...   print(row)

(None, None, None)
(None, None, None)
     
 
 
 -----------------
 jak mamy cele to mozemy przypisać wartości
 
 >>> c.value = 'hello, world'
>>> print(c.value)
'hello, world'

>>> d.value = 3.14
>>> print(d.value)
3.14   
"""





"""
zapis / nadpisu plik

>>> wb = Workbook()
>>> wb.save('balances.xlsx')





zapis jako template
>>> wb = load_workbook('document.xlsx')
>>> wb.template = True
>>> wb.save('document_template.xltx')





moż na zapisac jako stream  np do transmisji po sieci Flask,django....

>>> from tempfile import NamedTemporaryFile
>>> from openpyxl import Workbook
>>> wb = Workbook()
>>> with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
"""





"""
otwarcie istniejącego wb


>>> from openpyxl import load_workbook
>>> wb = load_workbook(filename = 'empty_book.xlsx')
>>> sheet_ranges = wb['range names']
>>> print(sheet_ranges['D18'].value)


"""