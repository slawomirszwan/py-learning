from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from my_lib import extract_data_dict, select_columns__dict

from itertools import groupby
from collections import Counter


def next_line_file_gen(filename):
    """
    generator zwraca kolejną linię pliku
    """
    with open(filename, 'r') as file:
        for line in file:
            yield line.rstrip('\n')


def extract_data_dict(data_file):
    """
    zwraca dane z pliku txt
    """
    row_gen = next_line_file_gen(data_file)

    header_line = next(row_gen)
    column_names = header_line.split()
    # print(f"{column_names=}")

    row_dict_gen = (dict(zip(column_names, line.split())) for line in row_gen)

    data = {
        "file": data_file,
        "column_names": column_names,
        "data": list(row_dict_gen)
    }
    return data


def select_columns__dict(row, selected_columns):
    """
    Wybiera tylko określone kolumny
    :param row:
    :param selected_columns:
    :return:
    """
    return {key: value for key, value in row.items() if key in selected_columns}


def validate_pesel(pesel):
    """
    waliduje PESEL
    :param pesel: str
    :return: True/False
    """
    if not pesel.isdigit() or len(pesel) != 11:
        return False

    weights = [1, 3, 7, 9, 1, 3, 7, 9, 1, 3]

    checksum = 10 - sum(int(p) * w % 10 for p, w in zip(pesel[0:-1], weights)) % 10
    # print(f'{checksum=}  {pesel[-1]}')
    return checksum == int(pesel[-1])


def extract_data_form_file():
    """
    extract data form txt file
    :return:
    """
    lekarz = extract_data_dict("../data_files/lekarze.txt")
    pacjent = extract_data_dict("../data_files/pacjenci.txt")
    wizyta = extract_data_dict("../data_files/wizyty.txt")
    return lekarz, pacjent, wizyta


def pacjent_transorm_func(row):
    """
    transform pacjent
    :param row:
    :return:
    """
    new_row = row.copy()
    new_row["Data_urodzenia"] = datetime.strptime(row["Data_urodzenia"], "%Y-%m-%d")
    return new_row


def lekarz_transorm_func(row):
    """
    transform lekarz
    :param row:
    :return:
    """
    new_row = row.copy()
    new_row["Specjalnosc"] = str.capitalize(row["Specjalnosc"])
    new_row["Data_urodzenia"] = datetime.strptime(row["Data_urodzenia"], "%Y-%m-%d")
    return new_row


def wizyta_transorm_func(row):
    """
    transform wizyta
    :param row:
    :return:
    """
    new_row = row.copy()
    new_row["Data_wizyty"] = datetime.strptime(row["Data_wizyty"], "%Y-%m-%d")
    return new_row


# ----------------------------------------------------------------
def show_data_from_file(lekarz, pacjent, wizyta):
    """
    pokazuje dane po etl
    :param lekarz:
    :param pacjent:
    :param wizyta:
    :return:
    """
    print('-' * 20, 'lekarz_lista', sep='\n')
    # print(lekarz.get("column_names"))
    print(*lekarz.get("data"), sep="\n")

    print('-' * 20, 'pacjent_lista', sep='\n')
    # print(pacjent.get("column_names"))
    print(*pacjent.get("data"), sep="\n")

    print('-' * 20, 'wizyta_lista', sep='\n')
    # print(wizyta.get("column_names"))
    print(*wizyta.get("data"), sep="\n")


def lekarze_posortowanie_po_nazwisku_reverse(data):
    """
    lekarze_posortowanie_po_nazwisku_reverse
    :param data:
    :return:
    """
    sorted_data = sorted(data, key=lambda x: x['Nazwisko'], reverse=True)
    print('-' * 20, 'lekarze_posortowanie_po_nazwisku_reverse', sep='\n')
    print(*sorted_data, sep="\n")


def lekarze_wybrane_kolumny(data, selected_columns):
    """
    lekarze_wybrane_kolumny
    :param data:
    :param selected_columns:
    :return:
    """
    filtered_columns = map(lambda r: select_columns__dict(r, selected_columns), data)
    print('-' * 20, 'lekarze_wybrane_kolumny', sep='\n')
    print(*filtered_columns, sep='\n')


def pacjenci_z_nieprawidłowym_pesel(data):
    """
    pacjenci_z_nieprawidłowym_pesel
    :param data:
    :return:
    """
    filtered_row = filter(lambda r: not validate_pesel(r['PESEL']), data)
    # print(*filtered_row, sep="\n")

    return filtered_row



def get_pacjenci_liczba_wizyt(data):
    """
    przy pomocy Counter
    :param data:
    :return:
    """
    #grupowanie wizyty po Id_pacjenta
    key_func = lambda row: row["Id_pacjenta"]
    grouped_data = groupby(sorted(data, key=key_func), key=key_func)

    # użycie Counter do zliczenia występowania elementów w każdej grupie
    counter = Counter( {  key: len(list(group)) for key, group in grouped_data})

    # print(counter)
    # # Counter({'100': 12, '311': 10, '319': 7, '164': 6, '273': 6, '255': 5, '317': 5, '181': 4, '205': 4, '231': 4, '249': 4, '262': 4, '286': 4, '110': 3, '121': 3, '124': 3, '135': 3, '161': 3, '170': 3, '173': 3, '191': 3, '200': 3, '208': 3, '209': 3, '213': 3, '226': 3, '227': 3, '233': 3, '238': 3, '241': 3, '246': 3, '257': 3, '272': 3, '291': 3, '299': 3, '301': 3, '309': 3, '314': 3, '316': 3, '401': 3, '111': 2, '122': 2, '147': 2, '155': 2, '160': 2, '165': 2, '166': 2, '172': 2, '180': 2, '184': 2, '193': 2, '198': 2, '204': 2, '207': 2, '212': 2, '236': 2, '245': 2, '252': 2, '258': 2, '266': 2, '271': 2, '284': 2, '288': 2, '292': 2, '303': 2, '313': 2, '404': 2, '405': 2, '409': 2, '415': 2, '416': 2, '163': 1, '189': 1, '192': 1, '197': 1, '216': 1, '235': 1, '243': 1, '244': 1, '247': 1, '248': 1, '253': 1, '263': 1, '264': 1, '274': 1, '276': 1, '283': 1, '285': 1, '290': 1, '295': 1, '296': 1, '300': 1, '306': 1, '307': 1, '308': 1, '403': 1, '408': 1, '411': 1, '412': 1, '418': 1, '420': 1})
    #
    # print(counter.most_common())
    # # [('100', 12), ('311', 10), ('319', 7), ('164', 6), ('273', 6), ('255', 5), ('317', 5), ('181', 4), ('205', 4), ('231', 4), ('249', 4), ('262', 4), ('286', 4), ('110', 3), ('121', 3), ('124', 3), ('135', 3), ('161', 3), ('170', 3), ('173', 3), ('191', 3), ('200', 3), ('208', 3), ('209', 3), ('213', 3), ('226', 3), ('227', 3), ('233', 3), ('238', 3), ('241', 3), ('246', 3), ('257', 3), ('272', 3), ('291', 3), ('299', 3), ('301', 3), ('309', 3), ('314', 3), ('316', 3), ('401', 3), ('111', 2), ('122', 2), ('147', 2), ('155', 2), ('160', 2), ('165', 2), ('166', 2), ('172', 2), ('180', 2), ('184', 2), ('193', 2), ('198', 2), ('204', 2), ('207', 2), ('212', 2), ('236', 2), ('245', 2), ('252', 2), ('258', 2), ('266', 2), ('271', 2), ('284', 2), ('288', 2), ('292', 2), ('303', 2), ('313', 2), ('404', 2), ('405', 2), ('409', 2), ('415', 2), ('416', 2), ('163', 1), ('189', 1), ('192', 1), ('197', 1), ('216', 1), ('235', 1), ('243', 1), ('244', 1), ('247', 1), ('248', 1), ('253', 1), ('263', 1), ('264', 1), ('274', 1), ('276', 1), ('283', 1), ('285', 1), ('290', 1), ('295', 1), ('296', 1), ('300', 1), ('306', 1), ('307', 1), ('308', 1), ('403', 1), ('408', 1), ('411', 1), ('412', 1), ('418', 1), ('420', 1)]
    # # list
    #
    # print(f"chempion: {counter.most_common()[0]=}")
    # # chempion: counter.most_common()[0]=('100', 12)  #first
    #
    # print(f"podium: {counter.most_common()[:3]=}")
    # # podium: counter.most_common()[:3]=[('100', 12), ('311', 10), ('319', 7)]
    return counter.most_common()


"""
zadanie 1:
lista pacjentów i liczbę ich wizyt   
pokolorowanych
3 pacjentów najczeciej korzystających z wizyt
1 miejsce czerwony
2 drugie pomarańczowy
3 trzecie zielony
"""
def pacjenci_ilosc_wizyt_excel(data, column_names, pacjenci_liczba_wizyt):
    """
    generowanie pliku excel
    :param data:
    :return:
    """

    wb = Workbook()

    ws = wb.active
    ws.title = "Pacjenci"

    # Kolor wiersza
    fills = [
        PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
        PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    ]

    #get trzech którzy najliczniej
    id_pacjents_for_colored = [ row[0] for row in pacjenci_liczba_wizyt[:3] ]

    ws.append(column_names)

    """
    kolor tylko dla komórki a nie całego wiersza 
    """
    for row in data:
        id_pacjenta = row['Id_pacjenta']
        has_color = id_pacjenta in id_pacjents_for_colored
        if has_color:
            # ustalanie koloru
            index_in_fills = id_pacjents_for_colored.index(id_pacjenta)
            fill_for_cell = fills[index_in_fills]

        cells_values = list(row.values())
        next_row = ws.max_row + 1
        next_column = 1
        for cell in cells_values:
            ws.cell(row=next_row, column=next_column).value = cell  # Wartość komórki

            if has_color:
                ws.cell(row=next_row, column=next_column).fill = fill_for_cell # kolor

            next_column +=1

    wb.save("Pacjenci_liczba_wizyt.xlsx")


# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

def pacjenci_blad_pesel_excel(data, column_names):
    """
    generowanie pliku excel
    :param data:
    :return:
    """

    wb = Workbook()

    ws = wb.active
    ws.title = "Pacjenci z nieprawidłowym PESEL"

    ws.append(column_names)

    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in data:

        has_color = not validate_pesel(row["PESEL"])

        cells_values = list(row.values())
        next_row = ws.max_row + 1
        next_column = 1
        for cell in cells_values:
            ws.cell(row=next_row, column=next_column).value = cell

            if has_color and next_column==4:
                ws.cell(row=next_row, column=next_column).fill = fill_red

            next_column += 1

    wb.save("Pacjenci_z_nieprawidłowym_PESEL.xlsx")

def main():
    lekarz, pacjent, wizyta = extract_data_form_file()
    # lekarz['data'] = [lekarz_transorm_func(row) for row in lekarz.get("data")]
    lekarz['data'] = list(map(lekarz_transorm_func, lekarz.get("data")))

    # pacjent['data'] = [pacjent_transorm_func(row) for row in pacjent.get("data")]
    pacjent['data'] = list(map(pacjent_transorm_func, pacjent.get("data")))

    # wizyta['data'] = [wizyta_transorm_func(row) for row in wizyta.get("data")]
    wizyta['data'] = list(map(wizyta_transorm_func, wizyta.get("data")))

    # show_data_from_file(lekarz, pacjent, wizyta)

    # lekarze_posortowanie_po_nazwisku_reverse(lekarz.get('data'))
    #
    # lekarze_wybrane_kolumny(lekarz.get("data"), ['Imie', 'Nazwisko'])


    # --------------------------
    pacjenci_liczba_wizyt = get_pacjenci_liczba_wizyt(wizyta.get('data'))
    # print(f"{pacjenci_liczba_wizyt=}")

    dict_id_count = {item[0]: item[1] for item in pacjenci_liczba_wizyt}

    data_to_merge = list(pacjent.get("data"))
    for item in data_to_merge:
        item["Liczba_wizyt"] = dict_id_count.get(item["Id_pacjenta"], 0)
    # print(*data_to_merga,sep="\n")

    new_column_names = pacjent.get("column_names") + ["liczba_wizyt"]
    # new_column_names = list(data_to_merge[0].keys())

    pacjenci_ilosc_wizyt_excel(data_to_merge, new_column_names, pacjenci_liczba_wizyt)


    # pacjenci_excel(pacjent.get("data"), pacjent.get("column_names"), pacjenci_liczba_wizyt)

    # --------------------------


    # data_table = pacjenci_z_nieprawidłowym_pesel(pacjent.get('data'))
    # print(*data_table, sep="\n")
    # print(len(list(data_table)))

    pacjenci_blad_pesel_excel(pacjent.get('data'), pacjent.get("column_names"))


if __name__ == "__main__":
    main()
