from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from collections import Counter

def extract_data_from_file(filename, transform_func):
    """
    zwraca dane z pliku txt
    pierwszy wiersz kolumny
    """
    with open(filename, 'r') as file:
        header_line = file.readline()
        column_names = header_line.split()

        data = []
        for line in file:
            values = line.split()
            row = {column_names[i]:v for i, v in enumerate(values)}
            row = transform_func(row)
            data.append(row)
        return data

def transform_lekarze(row):
    """
    transformuje dane
    map do nowej postaci
    :param row:
    :return:
    """
    row["Specjalnosc"] = row["Specjalnosc"].capitalize()
    row["Data_urodzenia"] = datetime.strptime(row["Data_urodzenia"], "%Y-%m-%d")
    return row

def transform_pacjenci(row):
    row["Data_urodzenia"] = datetime.strptime(row["Data_urodzenia"], "%Y-%m-%d")
    return row

def transform_wizyty(row):
    row["Data_wizyty"] = datetime.strptime(row["Data_wizyty"], "%Y-%m-%d")
    return row

# def group_by(data, key_func):
#     """
#     grupowanie
#     group_by(wizyty, lambda row: row['Id_pacjenta'])
#     group_by(pacjenci, lambda row: row['Imie'])
#     group_by(pacjenci, lambda row: f"{row['Imie'] row['Nazwisko']")
#     group_by(wizyty, lambda row: row['Data_wizyty'].year )  grupowanie wizyt po roku w którym odbyła się wizyta
#     group_by(wizyty, lambda row: row['Data_wizyty'].strftime('%A'))  grupowanie wizyt po dniu tygodnia w którym odbyła się wizyta
#     group_by(wizyty, lambda row: f"{row['Data_wizyty'].year} {row['Data_wizyty'].month}")   grupowanie wizyt po roku i miesiącu w którym odbyła się wizyta (podsumowanie miesiąca w danym roku)
#     :param data:
#     :param key_func:
#     :return: dictionary   key , group
#     """
#     groups_dict = {}
#     for row in data:
#         # grouping key
#         key = key_func(row)
#         #actual group or []
#         group = groups_dict.get(key, [])
#         #new value group list
#         group.append(row)
#         groups_dict[key] = group
#     return groups_dict


def validate_pesel(pesel):
    """
    waliduje PESEL
    :param pesel: str
    :return: True/False
    """
    if not pesel.isdigit() or len(pesel) != 11:
        return False

    weights = [1, 3, 7, 9, 1, 3, 7, 9, 1, 3]

    checksum = 10 - sum(int(p) * w % 10 for p, w in zip(pesel[0:-1], weights)) % 10
    # print(f'{checksum=}  {pesel[-1]}')
    return checksum == int(pesel[-1])


##########################################################
def main():
    lekarze = extract_data_from_file("../data_files/lekarze.txt", transform_lekarze)
    # print(*lekarze, sep="\n")

    pacjenci = extract_data_from_file("../data_files/pacjenci.txt", transform_pacjenci)
    # print(*pacjenci, sep="\n")

    wizyty = extract_data_from_file("../data_files/wizyty.txt", transform_wizyty)
    # print(*wizyty, sep="\n")



    #######################################################################################
    counter_wizyt_pacjenta = Counter(wizyta['Id_pacjenta'] for wizyta in wizyty)
    # print(counter_wizyt_pacjenta)

    ilosc_wizyt_list = counter_wizyt_pacjenta.most_common()
    # print(counter_wizyt_pacjenta.most_common())
    # print(*counter_wizyt_pacjenta.most_common(), sep="\n")
    # print(ilosc_wizyt_list)
    ilosc_wizyt_dict = dict(ilosc_wizyt_list)
    # print(ilosc_wizyt_dict)
    # """
    # {'100': 12, '311': 10, '319': 7, '164': 6,....
    # """

    # dodajemy kolumnę liczba_wizyt
    pacjenci_with_count_wizyt = pacjenci.copy()
    for pacjent in pacjenci_with_count_wizyt:
        id_pacjenta = pacjent['Id_pacjenta']
        pacjent['Liczba_wizyt'] = ilosc_wizyt_dict[id_pacjenta] if id_pacjenta in ilosc_wizyt_dict else 0
    # print(*pacjenci_with_count_wizyt, sep="\n")

    column_names = ['Id_pacjenta', 'Nazwisko', 'Imie', 'PESEL', 'Data_urodzenia', 'Liczba_wizyt']

    wb = Workbook()

    ws = wb.active
    ws.title = "Pacjenci oraz liczba wizyt"

    # Kolor wiersza
    fills = [
        PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
        PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    ]

    #get trzech którzy najliczniej
    id_pacjents_for_colored = [ row[0] for row in ilosc_wizyt_list[:3] ]

    ws.append(column_names)

    """
    kolor tylko dla komórki a nie całego wiersza 
    """
    for row in pacjenci_with_count_wizyt:
        id_pacjenta = row['Id_pacjenta']
        has_color = id_pacjenta in id_pacjents_for_colored
        if has_color:
            # ustalanie koloru
            index_in_fills = id_pacjents_for_colored.index(id_pacjenta)
            fill_for_cell = fills[index_in_fills]

        cells_values = list(row.values())
        next_row = ws.max_row + 1
        next_column = 1
        for cell in cells_values:
            ws.cell(row=next_row, column=next_column).value = cell  # Wartość komórki

            if has_color:
                ws.cell(row=next_row, column=next_column).fill = fill_for_cell # kolor

            next_column +=1

    wb.save("Pacjenci_z_liczba_wizyt.xlsx")

    ###############################################################################

    wb = Workbook()

    ws = wb.active
    ws.title = "Pacjenci z nieprawidłowym PESEL"

    ws.append(column_names)

    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in pacjenci_with_count_wizyt:

        has_color = not validate_pesel(row["PESEL"])

        cells_values = list(row.values())
        next_row = ws.max_row + 1
        next_column = 1
        for cell in cells_values:
            ws.cell(row=next_row, column=next_column).value = cell

            if has_color and next_column==4:
                ws.cell(row=next_row, column=next_column).fill = fill_red

            next_column += 1

    wb.save("Pacjenci_z_nieprawidłowym_PESEL.xlsx")

if __name__ == "__main__":
    main()

